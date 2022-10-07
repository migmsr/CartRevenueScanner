from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from time import sleep
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os

# excel manager
excelFile = 'Data.xlsx'

wb = Workbook()
wb = load_workbook(filename=excelFile)
sheetName = 'Plan1'
    
try:
    ws = wb.get_sheet_by_name(sheetName)

except:
    ws = wb.create_sheet(sheetName)

if (os.path.exists(excelFile)):
    os.remove(excelFile)
    wb.save(excelFile)
wb.save(excelFile)

####setting up headless option

url = 'https://www.cnj.jus.br/corregedoria/justica_aberta/?'

user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36'

# options = webdriver.ChromeOptions()
# options.headless = True
# options.add_argument(f'user-agent={user_agent}')
# options.add_argument("--window-size=1920,1080")
# options.add_argument('--ignore-certificate-errors')
# options.add_argument('--allow-running-insecure-content')
# options.add_argument("--disable-extensions")
# options.add_argument("--proxy-server='direct://'")
# options.add_argument("--proxy-bypass-list=*")
# options.add_argument("--start-maximized")
# options.add_argument('--disable-gpu')
# options.add_argument('--disable-dev-shm-usage')
# options.add_argument('--no-sandbox')
# browser = webdriver.Chrome(executable_path="chromedriver.exe", options=options)

###########################################

browser = webdriver.Edge()

browser.get(url)

browser.maximize_window()

action = webdriver.ActionChains(browser)

##############

first_click_xpath = '/html/body/div[2]/div[5]/fieldset/table/tbody/tr[1]/td[2]/a[1]'

uf_aux_xpath = '/html/body/div[2]/div[5]/fieldset/map'

uf_name_xpath = '/html/body/div[2]/div[5]/fieldset/table/tbody/tr/td[2]/form[2]/div/div/table/tbody/tr[1]/td[1]/strong'

select_cidades_xpath = '/html/body/div[2]/div[5]/fieldset/table/tbody/tr/td[2]/form[2]/div/div/table/tbody/tr[1]/td[2]/select'

pesquisar_bt_xpath = '/html/body/div[2]/div[5]/fieldset/table/tbody/tr/td[2]/form[2]/div/div/table/tbody/tr[2]/td/button[1]'

lista_cartorios_xpath = '/html/body/div[2]/div[5]/fieldset/div/table/tbody'

lista_info_xpath = '/html/body/div[2]/div[5]/fieldset[6]/table[2]/tbody'

##############

first_click = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, first_click_xpath))
)

first_click.click()

uf_aux = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, uf_aux_xpath))
)

uf_list = ('SP', 'BA', 'SC', 'RS', 'PR', 'RJ', 'ES', 'MG', 'SE', 'AL', 'PE', 
'PB', 'RN', 'CE', 'PI', 'MA', 'TO', 'GO', 'DF', 'MS', 'MT', 'RO', 'PA', 'RR', 'AM', 'AC', 'AP')

sleep(0.5)

for uf in uf_list:

    ws.append(['ESTADO: ', uf, '/////////', '/////////', '/////////', '/////////', '/////////'])

    if (os.path.exists(excelFile)):
        os.remove(excelFile)
        wb.save(excelFile)
    wb.save(excelFile)

    #executa script para selecionar o estado no mapa ao invés de clicar no mapa
    script = "pesquisaServentiasExtra('" + uf + "')"

    browser.execute_script(script)
    #

    #nome da uf atual para debug e controle
    uf_name = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, uf_name_xpath))
    )

    select_cidades = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, select_cidades_xpath))
    )

    cidades = select_cidades.find_elements(By.XPATH, '*')

    #transforma o elemento em um tipo select
    select_cidades_obj = Select(select_cidades)

    print("Estado atual: " + uf_name.get_attribute("innerText"))

    sleep(0.5)

    n_cidades = len(cidades)

    lista_ids_cidades = []
    lista_ids_cidades.clear()

    for i_cidade in range(1, n_cidades): #deve ignorar o primeiro elemento

        try:

            select_cidades = browser.find_element(By.XPATH, select_cidades_xpath)

        except:

            uf_aux = WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.XPATH, uf_aux_xpath))
            )
            #executa script para selecionar o estado no mapa ao invés de clicar no mapa
            script = "pesquisaServentiasExtra('" + uf + "')"

            browser.execute_script(script)
            #
            select_cidades = WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.XPATH, select_cidades_xpath))
            )

        cidades = select_cidades.find_elements(By.XPATH, '*')

        #transforma o elemento em um tipo select
        select_cidades_obj = Select(select_cidades)
       
        for cidade in cidades[1:]:

            id_cidade = cidade.get_attribute("innerText")    
            if(id_cidade in lista_ids_cidades):
                pass
            else:
                lista_ids_cidades.append(id_cidade)

                #lógica para selecionar as cidades
                print('Cidade: ' + cidade.get_attribute("innerText"))

                ws.append(['Cidade: ', cidade.get_attribute("innerText")])

                if (os.path.exists(excelFile)):
                    os.remove(excelFile)
                    wb.save(excelFile)
                wb.save(excelFile)

                select_cidades_obj.select_by_visible_text(cidade.get_attribute("innerText"))

                #espera a opção ser selecionada para depois clicar
                while(select_cidades_obj.first_selected_option.text != cidade.get_attribute("innerText")):
                    pass

                pesquisar_bt = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, pesquisar_bt_xpath))
                )

                pesquisar_bt.click()

                
                #cria lista dos cartórios e conta quantos têm
                lista_cartorios = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, lista_cartorios_xpath))
                )

                lista_cartorios = lista_cartorios.find_elements(By.XPATH, '*')
                
                lista_ids_cartorios = []
                lista_ids_cartorios.clear()

                n_cartorios = len(lista_cartorios)

                for i_cartorio in range(n_cartorios):

                    #cria lista dos cartórios e passa por cada um
                    lista_cartorios = WebDriverWait(browser, 10).until(
                    EC.presence_of_element_located((By.XPATH, lista_cartorios_xpath))
                    )

                    lista_cartorios = lista_cartorios.find_elements(By.XPATH, '*')
                    
                    for cartorio in lista_cartorios:
                        
                        id_cartorio = cartorio.get_attribute("id")

                        if(id_cartorio in lista_ids_cartorios):
                            pass
                        else:
                            lista_ids_cartorios.append(id_cartorio)
                            #acessa dados do responsavel
                            responsavel = cartorio.find_element(By.CSS_SELECTOR, 'tbody').find_elements(By.XPATH, '*')[1].find_elements(By.XPATH, '*')[1].get_attribute("innerText")
                            print(responsavel)

                            if(responsavel != ""):

                                ws.append(['Responsável', responsavel])

                                if (os.path.exists(excelFile)):
                                    os.remove(excelFile)
                                    wb.save(excelFile)
                                wb.save(excelFile)
                                
                                situacao = cartorio.find_elements(By.XPATH, '*')[2].get_attribute("innerText")
                                print(situacao)

                                ws.append(['Situação: ', situacao])

                                if (os.path.exists(excelFile)):
                                    os.remove(excelFile)
                                    wb.save(excelFile)
                                wb.save(excelFile)

                                #acessa a página com os dados do cartorio
                                script = "wiOpen('?d=consulta_extra&a=consulta_extra&f=formDadosServentiaExtra&SEQ_DADOS_SERVENTIA=" + id_cartorio + "');"
                                #ao inves de clicar, executa o mesmo script que a pagina executa para acessar os dados
                                browser.execute_script(script)
                                
                                lista_info = WebDriverWait(browser, 10).until(
                                EC.presence_of_element_located((By.XPATH, lista_info_xpath ))
                                )

                                list_info = lista_info.find_elements(By.XPATH, '*')

                                for info in list_info[1:-1]:

                                    info_vector = info.get_attribute("innerText").split('\n')
                                    periodo = info_vector[0]
                                    arrecadacao = info_vector[3]
                                    ano = periodo[-5:]
                                    ano = ano[:4]
                                    if(ano == '2021'):
                                        print(periodo)
                                        print(arrecadacao)

                                        ws.append([periodo, arrecadacao])

                                        if (os.path.exists(excelFile)):
                                            os.remove(excelFile)
                                            wb.save(excelFile)
                                        wb.save(excelFile)

                                #volta para a página anterior
                                browser.back()
                                break

                            else:
                                print("Responsável ausente. Cartório inativo.")

                                ws.append(['Responsável ausente. Cartório inativo.'])

                                if (os.path.exists(excelFile)):
                                    os.remove(excelFile)
                                    wb.save(excelFile)
                                wb.save(excelFile)

                            ws.append(['///////////', '///////////', '///////////', '///////////', '///////////'])

                            if (os.path.exists(excelFile)):
                                os.remove(excelFile)
                                wb.save(excelFile)
                            wb.save(excelFile)
                            
            
                browser.back()
                break
        

        # for cartorio in lista_cartorios:
            
        #     #acessa dados do responsavel
        #     responsavel = cartorio.find_element(By.CSS_SELECTOR, 'tbody').find_elements(By.XPATH, '*')[1].find_elements(By.XPATH, '*')[1].get_attribute("innerText")
        #     print(responsavel)

        #     if(responsavel != ""):
                
        #         situacao = cartorio.find_elements(By.XPATH, '*')[2].get_attribute("innerText")
        #         print(situacao)
        #         #acessa a página com os dados do cartorio
        #         id_cartorio = cartorio.get_attribute("id")
        #         script = "wiOpen('?d=consulta_extra&a=consulta_extra&f=formDadosServentiaExtra&SEQ_DADOS_SERVENTIA=" + id_cartorio + "');"
        #         #ao inves de clicar, executa o mesmo script que a pagina executa para acessar os dados
        #         browser.execute_script(script)
                
        #         lista_info = WebDriverWait(browser, 10).until(
        #         EC.presence_of_element_located((By.XPATH, lista_info_xpath ))
        #         )

        #         list_info = lista_info.find_elements(By.XPATH, '*')

        #         for info in list_info[1:-1]:

        #             info_vector = info.get_attribute("innerText").split('\n')
        #             periodo = info_vector[0]
        #             arrecadacao = info_vector[3]
        #             ano = periodo[-5:]
        #             ano = ano[:4]
        #             if(ano == '2021'):
        #                 print(periodo)
        #                 print(arrecadacao)

        #         #volta para a página anterior
        #         browser.back()

        #     else:
        #         print("Responsável ausente. Cartório inativo.")

            

#browser.close()

