from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import random
import requests
import os
import shutil

pid = os.getpid()

print('Process PID:', pid)

#os.startfile('Bot Controller\dist\BotController.exe')


def RandomDelays():
    vector = [1, 2, 3]
    
    delayRange = random.randint(1, 30)
    delay = 1
    for i in range(delayRange):
        delay += 0.1 
        
    vector[0] = delay
    
    delayRange = random.randint(1, 30)
    delay = 0.5
    for i in range(delayRange):
        delay += 0.03
        
    vector[1] = delay
    
    delayRange = random.randint(1, 30)
    delay = 2
    for i in range(delayRange):
        delay += 0.06
        
    vector[2] = delay
    
    return vector

#excel manager

excelFile = 'Data.xlsx'
bkpFile = 'Data - bkp.xlsx'
wb = Workbook()
wb = load_workbook(filename = excelFile)
sheetNumber = '1'

shutil.copyfile(excelFile, bkpFile)

while(True):
    
    try:
        ws = wb.get_sheet_by_name(sheetNumber)
        
    except:
        ws = wb.create_sheet(sheetNumber)
        wb.save(excelFile)
        break
    
    sheetNumber = str(int(sheetNumber) + 1)
        
    
#####################################


#whatsapp manager

from twilio.rest import Client

account_sid = 'AC9b7c6f59061748c95823963d7f930796'

auth_token = '0861b0e54c6d71f72061499dd83012f9'

client = Client(account_sid, auth_token)

from_whatsapp_number = 'whatsapp:+14155238886'

to_whatsapp_number = 'whatsapp:+553899135151'

####

####setting up headless option

url = 'https://blaze.com/pt/games/double'

user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36'

options = webdriver.ChromeOptions()
options.headless = True
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument("--start-maximized")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')
browser = webdriver.Chrome(executable_path="chromedriver.exe", options=options)

###########################################

#browser = webdriver.Chrome()

time.sleep(1)

browser.get('https://blaze.com/pt/games/double')

time.sleep(1)

browser.save_screenshot('screenshot.png')

browser.maximize_window()

lastRound_xpath = '/html/body/div[1]/main/div[1]/div[4]/div/div[1]/div/div/div[1]/div[2]/div[2]/div/div[1]/div[1]/div/div'

i = 2

############################################################
#auto login

minStreak = 5
minDisparity = 5
minRounds = 200
shouldTakeProfit = False

bet_amount = 2

login = 'migmsr@hotmail.com'

password = 'Miguel.99062535'

bet = -1

signin_bt_xpath = '/html/body/div[1]/main/div[1]/div[2]/div/div[2]/div/div[2]/div/div/div[1]/a'

login_input_xpath = '/html/body/div[1]/main/div[3]/div/div[2]/div[2]/form/div[1]/div/input'

password_input_xpath = '/html/body/div[1]/main/div[3]/div/div[2]/div[2]/form/div[2]/div/input'

submit_button_xpath = '/html/body/div[1]/main/div[3]/div/div[2]/div[2]/form/div[4]/button'

signin_bt = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, signin_bt_xpath))
)

signin_bt.click()

login_input = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, login_input_xpath))
)

login_input.send_keys(login)

time.sleep(2)

password_input = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, password_input_xpath))
)

password_input.send_keys(password)

time.sleep(2)

submit_button = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, submit_button_xpath))
)

submit_button.click()

time.sleep(2)

# while bet_amount == 0:
    
#     bet_amount = float(input('Sign in and type bet amount to continue: '))
#     pass

###############################################################

bet_button_xpath = '/html/body/div[1]/main/div[1]/div[4]/div/div[1]/div/div/div[1]/div[1]/div[1]/div[3]/button'

lastRound_xpath = '/html/body/div[1]/main/div[1]/div[4]/div/div[1]/div/div/div[1]/div[2]/div[2]/div/div[1]/div[1]/div/div'

black_button_xpath = '/html/body/div[1]/main/div[1]/div[4]/div/div[1]/div/div/div[1]/div[1]/div[1]/div[2]/div[2]/div/div[3]/div'

red_button_xpath = '/html/body/div[1]/main/div[1]/div[4]/div/div[1]/div/div/div[1]/div[1]/div[1]/div[2]/div[2]/div/div[1]/div'

amount_field_xpath = '/html/body/div[1]/main/div[1]/div[4]/div/div[1]/div/div/div[1]/div[1]/div[1]/div[2]/div[1]/div/div[1]/input'

wallet_xpath = '/html/body/div[1]/main/div[1]/div[2]/div/div[2]/div/div[2]/div/div[3]/div/a'

color_number_xpath = '/html/body/div[1]/main/div[1]/div[4]/div/div[1]/div/div/div[1]/div[2]/div[2]/div/div[1]/div[1]/div/div/div'

bet_button = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, bet_button_xpath))
)

lastRound = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, lastRound_xpath))
)

try:
    color_number = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, color_number_xpath))
    )
except:
    pass
    
amount_field = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, amount_field_xpath))
)

black_button = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, black_button_xpath))
)

red_button = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, red_button_xpath))
)

wallet = WebDriverWait(browser, 10).until(
EC.presence_of_element_located((By.XPATH, wallet_xpath))
)


#wallet field conversion to str and int

size = len(str(wallet.accessible_name))
wallet_str = str(wallet.accessible_name[:size - 2])
wallet_str = wallet_str[2:]
print('Current Wallet: ' + wallet_str)


maxWallet = float(wallet_str)
current_wallet_str = wallet_str
drawdown = 0
current_drawdown = 0

realWallet = maxWallet/0.2 - maxWallet

#################################################################

try:
    lastRound = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, lastRound_xpath))
    )
finally:
    
    start_time = time.time()
    
    #bet strategy
    
    n_rounds = 0
    
    red = 0
    black = 0
    white = 0
    bet = -1
    win = 0
    loss = 0
    streak = 0
    
    ###
    
    now = datetime.now()
    current_time = now.strftime('%H:%M:%S')
    print(current_time)
    print(lastRound.get_attribute("class")) #returns the color
    print(lastRound.id) #returns id of element
    
    #write data to excel
    
    timestamp = str(current_time)
    
    color_number_txt = '15'
    
    if(str(lastRound.get_attribute("class")) == 'sm-box red'):
        roundResult = 0
        red += 1
        ##########
        color_number = browser.find_element('xpath', color_number_xpath)
        color_number_txt = color_number.get_attribute('innerText')
        ##########
    elif(str(lastRound.get_attribute("class")) == 'sm-box black'):
        roundResult = 1
        black += 1
        ##########
        color_number = browser.find_element('xpath', color_number_xpath)
        color_number_txt = color_number.get_attribute('innerText')
        ##########
    elif(str(lastRound.get_attribute("class")) == 'sm-box white'):
        roundResult = 2
        white += 1
    
    roundId = str(lastRound.id)
    
    try:
        ping = 'Ping: ' + str(requests.get("https://blaze.com/pt/games/double").elapsed.total_seconds())
    except:   
        ping = 'Ping: error'
    
    
    dataList = [roundResult, timestamp, roundId, color_number_txt, ping]
    print(color_number_txt)
    print(ping)
        
    ws.append(dataList)
    if(os.path.exists(excelFile)):
        os.remove(excelFile)
        wb.save(excelFile)
    wb.save(excelFile)

    #
    

try:
      whatsapp_message = 'Bot started.\n'
      pid_str = 'Process PID: ' + str(pid) + '\n'
      whatsapp_message += pid_str
      whatsapp_message += 'Rules:\n'
      whatsapp_message += 'Min Disparity: ' + str(minDisparity) + '\n'
      whatsapp_message += 'Min Streak: ' + str(minStreak) + '\n'
      whatsapp_message += 'Min Rounds: ' + str(minRounds) + '\n'
      whatsapp_message += 'Bet Amount: ' + str(bet_amount) + '\n'
      whatsapp_message += 'Should Take Profit: ' + str(shouldTakeProfit) + '\n'
      print(whatsapp_message)
      client.messages.create(
        
        body=whatsapp_message,
        from_=from_whatsapp_number,
        to=to_whatsapp_number
        )
except:
    print("can't send whatsapp")
    
try:
      client.messages.create(
        
        media_url = ['screeshot.png'],
        from_=from_whatsapp_number,
        to=to_whatsapp_number
        )
except:
    print("can't send whatsapp")

running = 0

try:

    while True:
        
        # #
        # print(time.time() - start_time)
        # #
        
        #check if the last round id has changed (which means the round updated)
        currentRound_xpath = '/html/body/div[1]/main/div[1]/div[4]/div/div[1]/div/div/div[1]/div[2]/div[2]/div/div[1]/div[1]/div/div'
        
        try:
            currentRound = WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.XPATH, currentRound_xpath))
            )
                        
            if(currentRound.id != lastRound.id):
                rounds_interval = str(round(time.time() - start_time)) + ' seconds'
                start_time = time.time()
                lastRound = currentRound
                now = datetime.now()
                current_time = now.strftime('%H:%M:%S')
                i += 1
                print(current_time) #returns timestamp of round
                print(lastRound.get_attribute("class")) #returns the color
                print(lastRound.id) #returns id of element
                
                #Data for excel
                timestamp = str(current_time)
                
                color_number_txt = '15'
                
                if(str(lastRound.get_attribute("class")) == 'sm-box red'):
                    roundResult = 0
                    red += 1
                    ##########
                    color_number = browser.find_element('xpath', color_number_xpath)
                    color_number_txt = color_number.get_attribute('innerText')
                    ##########
                elif(str(lastRound.get_attribute("class")) == 'sm-box black'):
                    roundResult = 1
                    black += 1
                    ##########
                    color_number = browser.find_element('xpath', color_number_xpath)
                    color_number_txt = color_number.get_attribute('innerText')
                    ##########
                elif(str(lastRound.get_attribute("class")) == 'sm-box white'):
                    roundResult = 2
                    white += 1
                
                roundId = str(lastRound.id)
                
                try:
                    ping = 'Ping: ' + str(requests.get("https://blaze.com/pt/games/double").elapsed.total_seconds())
                except:   
                    ping = 'Ping: error'
                    
                dataList = [roundResult, timestamp, roundId, color_number_txt, ping, rounds_interval]
                print(color_number_txt)
                print(ping)
                print(rounds_interval)
                
                #
                
                #bet_strategy
                
                n_rounds += 1
                    
                disparity = red - black
                
                dataList.append('Disparity: ' + str(disparity))
                
                dataList.append('Streak: ' + str(streak))
                
                if(bet != -1):
                    if(bet == roundResult):
                        win += 1
                        dataList.append('win')
                        
                        # #############
                        # #comment this for real application
                        # temp = float(current_wallet_str)
                        # temp += bet_amount
                        # current_wallet_str = str(temp)
                        # #############
                        
                        #############
                        #comment this for real application
                        if(float(current_wallet_str) > maxWallet):
                            maxWallet = float(current_wallet_str)
                            n_rounds = 0
                            print('Take Profit')
                            dataList.append('Take Profit')
                            try:
                                whats_result_message = 'Take Profit'
                                client.messages.create(
                                    
                                    body=whats_result_message,
                                    from_=from_whatsapp_number,
                                    to=to_whatsapp_number
                                    )
                            except:
                                print("can't send whatsapp")
                        ############################
                            
                        #################
                        
                        try:
                            whats_result_message = 'Result: Win'
                            client.messages.create(
                                
                                body=whats_result_message,
                                from_=from_whatsapp_number,
                                to=to_whatsapp_number
                                )
                        except:
                            print("can't send whatsapp")
                            
                        print("win")
                            
                    else:
                        loss += 1
                        dataList.append('loss')
                        
                        # #############
                        # #comment this for real application
                        # temp = float(current_wallet_str)
                        # temp -= bet_amount
                        # current_wallet_str = str(temp)
                        # #############
                        
                        try:
                            whats_result_message = 'Result: Loss'
                            client.messages.create(
                                
                                body=whats_result_message,
                                from_=from_whatsapp_number,
                                to=to_whatsapp_number
                                )
                            
                        except:
                            print("can't send whatsapp")
                            
                        print("loss")
                    bet = -1
                
                if(n_rounds >= minRounds):
                    
                    if(disparity > 0): #count number of non-black
                        if(roundResult != 1):
                            streak += 1
                        else:
                            streak = 0
                        
                        if(streak >= minStreak and disparity >= minDisparity):
                            #make a bet with random delay
                            bet = 1
                            print('############################')
                            print('Disparity: ', disparity)
                            print('Streak: ', streak)
                            print('Bet on black')
                            dataList.append('Bet on black')
                            
                            try:
                                whats_bet_message = 'Bet on black'
                                client.messages.create(
                                    
                                    body=whats_bet_message,
                                    from_=from_whatsapp_number,
                                    to=to_whatsapp_number
                                    )
                                
                            except:
                                print("can't send whatsapp")
                                
                            streak = 0
                            
                            if(bet_amount <= float(wallet_str)):
                                
                                
                                delay = RandomDelays()
                                
                                print('Delay: ' + str(delay[0]) + ' seconds')
                                time.sleep(delay[0])
                                #fill amount field
                                amount_field.clear()
                                amount_field.send_keys(bet_amount)
                                
                                while(bet_button.is_enabled() == False):
                                    pass
                               
                                print('Delay: ' + str(delay[1]) + ' seconds')
                                time.sleep(delay[1])
                                #click bet color
                                black_button.click()
                                
                                print('Delay: ' + str(delay[2]) + ' seconds')
                                time.sleep(delay[2])
                                #submit the bet
                                bet_button.click()
                                
                                print('Bet confirmed')
                                
                                try:
                                    whats_bet_message = 'Bet confirmed'
                                    client.messages.create(
                                        
                                        body=whats_bet_message,
                                        from_=from_whatsapp_number,
                                        to=to_whatsapp_number
                                        )
                                    
                                except:
                                    print("can't send whatsapp")
                                    
                            else:
                                print('Not enough funds.')
                                print('Recommended to review strategy or code.')
                                
                                try:
                                    whats_bet_message = 'Not enough funds\nRecommended to review strategy or code.'
                                    client.messages.create(
                                        
                                        body=whats_bet_message,
                                        from_=from_whatsapp_number,
                                        to=to_whatsapp_number
                                        )
                                    
                                except:
                                    print("can't send whatsapp")
                                    
                            
                    if(disparity < 0):
                        if(roundResult != 0):
                            streak += 1
                        else:
                            streak = 0
                        if(streak >= minStreak and disparity <= -1*minDisparity):
                            bet = 0
                            #make a bet with random delay
                            print('############################')
                            print('Disparity: ', disparity)
                            print('Streak: ', streak)
                            print('Bet on red')
                            
                            try:
                                whats_bet_message = 'Bet on red'
                                client.messages.create(
                                    
                                    body=whats_bet_message,
                                    from_=from_whatsapp_number,
                                    to=to_whatsapp_number
                                    )
                                
                            except:
                                print("can't send whatsapp")
                                
                            
                            dataList.append('Bet on red')
                            streak = 0
                            
                            if(bet_amount <= float(wallet_str)):
                            
                                delay = RandomDelays()
                                
                                print('Delay: ' + str(delay[0]) + ' seconds')
                                time.sleep(delay[0])
                                #fill amount field
                                amount_field.clear()
                                amount_field.send_keys(bet_amount)
                                
                                while(bet_button.is_enabled() == False):
                                    pass
                               
                                print('Delay: ' + str(delay[1]) + ' seconds')
                                time.sleep(delay[1])
                                #click bet color
                                red_button.click()
                                
                                print('Delay: ' + str(delay[2]) + ' seconds')
                                time.sleep(delay[2])
                                #submit the bet
                                bet_button.click()
                                
                                print('Bet confirmed')
                                
                                try:
                                    whats_bet_message = 'Bet confirmed'
                                    client.messages.create(
                                        
                                        body=whats_bet_message,
                                        from_=from_whatsapp_number,
                                        to=to_whatsapp_number
                                        )
                                    
                                except:
                                    print("can't send whatsapp")
                                    
                                    
                            else:
                                print('Not enough funds.')
                                print('Recommended to review strategy or code.')
                                
                                try:
                                    whats_bet_message = 'Not enough funds\nRecommended to review strategy or code.'
                                    client.messages.create(
                                        
                                        body=whats_bet_message,
                                        from_=from_whatsapp_number,
                                        to=to_whatsapp_number
                                        )
                                    
                                except:
                                    print("can't send whatsapp")
                                    
                    
                    ########
                #write data on excel
                
                ws.append(dataList)
                if(os.path.exists(excelFile)):
                    os.remove(excelFile)
                    wb.save(excelFile)
                wb.save(excelFile)
                
                ####
        
        
            ###################################
            #uncomment this for real application
            try:
                current_wallet = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, wallet_xpath))
                )
                size = len(str(current_wallet.accessible_name))
                current_wallet_str = str(current_wallet.accessible_name[:size - 2])
                current_wallet_str = current_wallet_str[2:]
            except:
                
                print('Error updating wallet')
                
                try:
                    whats_bet_message = 'Error updating wallet'
                    client.messages.create(
                        
                        body=whats_bet_message,
                        from_=from_whatsapp_number,
                        to=to_whatsapp_number
                        )
                    
                except:
                    print("can't send whatsapp")
            ####################################
            
            if(current_wallet_str != wallet_str):
                if(float(current_wallet_str) > maxWallet):
                    maxWallet = float(current_wallet_str)
                    
                    #uncomment this for real application w/ take profit
                    ################################
                    if(shouldTakeProfit):
                        n_rounds = 0
                        print('Take Profit')
                        dataList.append('Take Profit')
                        try:
                            whats_result_message = 'Take Profit'
                            client.messages.create(
                                
                                body=whats_result_message,
                                from_=from_whatsapp_number,
                                to=to_whatsapp_number
                                )
                        except:
                            print("can't send whatsapp")
                    ############################
                    
                else:
                    #drawdown always considering that you only have 20% of the real funds available on the app
                    #so the drawdown will consider the "real wallet", outside the app
                    #the objective is to always have drawdown > -20%
                    current_drawdown = ((float(current_wallet_str) - maxWallet)/(realWallet + maxWallet)) * 100
                    if(current_drawdown < drawdown):
                        drawdown = current_drawdown
                    print('Drawdown: ', round(drawdown), '%')
                
                wallet_str = current_wallet_str
                print('Current Wallet: R$', wallet_str)
                try:
                    whats_bet_message = 'Current Wallet: R$' + wallet_str
                    whats_bet_message += '\nDrawdown: ' + str(round(drawdown)) + '%'
                    client.messages.create(
                        
                        body=whats_bet_message,
                        from_=from_whatsapp_number,
                        to=to_whatsapp_number
                        )
                    
                except:
                    print("can't send whatsapp")
                
        except:
            
            browser.refresh()
            n_rounds = 0
            red = 0
            black = 0
            white = 0
            bet = -1
            streak = 0
            
            start_time = time.time()
             
            print('Refreshing browser...')
            ws.append(['Refreshing browser...'])
            if(os.path.exists(excelFile)):
                os.remove(excelFile)
                wb.save(excelFile)
            wb.save(excelFile)
            
            try:
                  whatsapp_message = 'Refreshing browser...'
                  client.messages.create(
                    
                    body=whatsapp_message,
                    from_=from_whatsapp_number,
                    to=to_whatsapp_number
                    )
                  
            except:
                print("can't send whatsapp")
            
            try:
            
                bet_button = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, bet_button_xpath))
                )
     
                lastRound = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, lastRound_xpath))
                )
                
                color_number = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, color_number_xpath))
                )
     
                amount_field = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, amount_field_xpath))
                )
     
                black_button = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, black_button_xpath))
                )
     
                red_button = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, red_button_xpath))
                )
     
                wallet = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, wallet_xpath))
                )
                
            except:
                pass
        pass
        
       #######################################
        if(time.time() - start_time >= 60):
           
           browser.refresh()
           n_rounds = 0
           red = 0
           black = 0
           white = 0
           bet = -1
           streak = 0
           
           
           try:
               ping = 'Blaze Ping: ' + str(requests.get("https://blaze.com/pt/games/double").elapsed.total_seconds())
           
           except:   
               ping = 'Blaze Ping: error'
               
           try:
               ping2 = 'Google Ping: ' + str(requests.get("https://www.google.com/").elapsed.total_seconds())
           
           except:   
               ping2 = 'Google Ping: error'
               
           print('Browser needing refresh. ', ping, ping2)
           
           ws.append([ping])
           ws.append([ping2])
           if(os.path.exists(excelFile)):
               os.remove(excelFile)
               wb.save(excelFile)
           wb.save(excelFile)
               
           try:
               whatsapp_message = 'Browser needing refresh. ' + ping + ping2
               client.messages.create(
                 
                 body=whatsapp_message,
                 from_=from_whatsapp_number,
                 to=to_whatsapp_number
                 )
                     
           except:
               print("can't send whatsapp")
               
           
           try:
           
               bet_button = WebDriverWait(browser, 10).until(
               EC.presence_of_element_located((By.XPATH, bet_button_xpath))
               )
    
               lastRound = WebDriverWait(browser, 10).until(
               EC.presence_of_element_located((By.XPATH, lastRound_xpath))
               )
    
               amount_field = WebDriverWait(browser, 10).until(
               EC.presence_of_element_located((By.XPATH, amount_field_xpath))
               )
    
               black_button = WebDriverWait(browser, 10).until(
               EC.presence_of_element_located((By.XPATH, black_button_xpath))
               )
    
               red_button = WebDriverWait(browser, 10).until(
               EC.presence_of_element_located((By.XPATH, red_button_xpath))
               )
    
               wallet = WebDriverWait(browser, 10).until(
               EC.presence_of_element_located((By.XPATH, wallet_xpath))
               )
               
           except:
               pass
               
           start_time = time.time()
            
           print('Server or connection is probably lagging.')
           print('Restarting rounds counts now...')
           ws.append(['Server or connection is probably lagging.'])
           ws.append(['Restarting rounds counts...'])
           if(os.path.exists(excelFile)):
               os.remove(excelFile)
               wb.save(excelFile)
           wb.save(excelFile)
           
           try:
                 whatsapp_message = 'Server or connection is probably lagging.\n'
                 whatsapp_message += "Restarting rounds count..."
                 whatsapp_message += ping
                 client.messages.create(
                   
                   body=whatsapp_message,
                   from_=from_whatsapp_number,
                   to=to_whatsapp_number
                   )
                 
           except:
               print("can't send whatsapp")
               
           
        
        
    
        
except BaseException as e:
    #get current system exeption
    e_type, e_value, e_traceback = sys.exc_info()
    
    print("Exception type : %s " %e_type.__name__)
    print("Exception message : %s" %e_value)
    print('Exception line: %s' %e_traceback.tb_lineno)
    #print("Stack trace : %s" %stack_trace)
    
    ws.append([str(e_type.__name__), str(e_traceback.tb_lineno)])
    #ws.append([str(e_value)])
    
    data = ['Red', red, 'Black', black, 'White', white]
    ws.append(data)
    
    n_bets = win + loss
    winrate = 0
    format_winrate = 0
    if(n_bets > 0):
        winrate = win/n_bets
        winrate = winrate*100
        format_winrate = "{:.2f}".format(winrate)
    data = ['Wins', win, 'Losses', loss, "Win Rate", str(winrate) + '%', 'Drawdown: ', str(round(drawdown)) + '%']
    ws.append(data)
    
    ws.append(['Done'])
    if(os.path.exists(excelFile)):
        os.remove(excelFile)
        wb.save(excelFile)
    wb.save(excelFile)

    try:
          whatsapp_message = 'Bot Ended\nException Type: ' + str(e_type.__name__) + '\n'
          whatsapp_message += 'Wins: ' + str(win) + '\nLosses: ' + str(loss) + '\nWin Rate: ' + str(winrate) + '%'
          whatsapp_message += '\nDrawdown: ' + str(round(drawdown)) + '%'
          client.messages.create(
            
            body=whatsapp_message,
            from_=from_whatsapp_number,
            to=to_whatsapp_number
            )
          
    except:
        print("can't send whatsapp")
          

#sm-box black = black
#sm-box red = red
#sm-box white = white


